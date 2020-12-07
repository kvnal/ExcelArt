import argparse
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import cv2 as cv


def RBG2HEX(RBG):
    return "%02x%02x%02x"%(RBG[0],RBG[1],RBG[2])

def draw(img_path,save):    
    img = cv.imread(img_path)
    img_rgb=cv.cvtColor(cv.resize(img,None, fx=1/2,fy=1/2),cv.COLOR_BGR2RGB)
    row,col=img_rgb.shape[:2]

    workbook=openpyxl.Workbook()
    sheet=workbook.create_sheet("canvas",0)
    sheet.cell(1,1,"ZOOM OUT!!")

    for i in range(1,row):
        for j in range(1,col):
            sheet.cell(i,j).fill=PatternFill(start_color=RBG2HEX(img_rgb[i,j]), fill_type = "solid")
            sheet.column_dimensions[get_column_letter(j)].width=3 #25px
    
    return workbook.save(save+".xlsx")

parser=argparse.ArgumentParser(description="Conver Image to Excel Art!")
parser.add_argument("-f","--file",required=True,help="Path of image file",metavar="",type=str)
parser.add_argument("-o","--output",required=True,help="Saved name",metavar="",type=str)


args=parser.parse_args()

if __name__ == "__main__":
    draw(args.file,args.output)
    pass